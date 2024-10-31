from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
import unicodedata


def criar_dicionario_do_mes(mes_nome):

    servidores = {}
    for linha in range(2, mes_nome.max_row + 1):
        nome_servidor = mes_nome[f"{colunas['Nome Servidor']}{linha}"].value
        siape = mes_nome[f"{colunas['Siape']}{linha}"].value
        lotacao = mes_nome[f"{colunas['Lotação']}{linha}"].value
        etr = mes_nome[f"{colunas['Equipe de Trabalho Remoto']}{linha}"].value
        modalidade = mes_nome[f"{colunas['Modalidade']}{linha}"].value
        regime_execucao = mes_nome[f"{colunas['Regime de Execução']}{linha}"].value

        if siape:
            servidores[siape] = {
                "Nome servidor": nome_servidor,
                "Lotação": lotacao,
                "Equipe de Trabalho Remoto": etr,
                "Modalidade": modalidade,
                "Regime de Execução": regime_execucao,
            }

    return servidores


def normalizar_texto_celula(texto):
    """
    Remove espaços em branco de uma célula, convertendo para maiúsculas,
    ordenando os itens em ordem alfabética quando separados por '/' e removendo acentos.
    """
    if texto is None:
        return None

    # * Removendo os acentos
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")

    # * Remove espaços, transforma em maiúsculas e ordena os itens
    itens = [item.strip().upper() for item in texto.split("/")]
    itens.sort()

    return "/".join(itens)


def comparar_meses(mes_atual, mes_anterior):
    """
    Compara os dados entre o mês atual e o mês anterior,
    retornando as diferenças encontradas.
    """
    alteracoes = []

    # * Percorre todos os servidores do mês atual
    for siape, dados_atual in mes_atual.items():
        if siape in mes_anterior:
            dados_anterior = mes_anterior[siape]

            # * Compara os campos
            for chave in dados_atual.keys():
                valor_atual = normalizar_texto_celula(dados_atual[chave])
                valor_anterior = normalizar_texto_celula(dados_anterior[chave])

                if valor_atual != valor_anterior:
                    alteracoes.append(
                        {
                            "Nome servidor": dados_atual["Nome servidor"],
                            "Siape": siape,
                            "Alteração": chave,
                            "De": valor_anterior,
                            "Para": valor_atual,
                        }
                    )

    return alteracoes


def cria_planilha_com_alteracoes(
    planilha, nome_planilha, alteracoes, cabecalho, mes_atual
):
    """
    Função para criar uma nova planilha no DataFrame apenas com as informações que foram alteradas.
    """
    # * Nome da planilha
    if nome_planilha in planilha.sheetnames:  # verifica se a planilha já existe
        nova_planilha = planilha[nome_planilha]
    else:  # Cria uma nova planilha
        mes_atual_planilha = meses[mes_atual].title

        # Obtém o índice dessa aba
        index_mes = planilha.sheetnames.index(mes_atual_planilha)
        # Cria a nova planilha logo após a do mês atual
        nova_planilha = planilha.create_sheet(title=nome_planilha, index=index_mes + 1)

    # * Adicionando o cabeçalho
    nova_planilha.append(cabecalho)

    # * Adiciona alterações
    for alteracao in alteracoes:
        nova_planilha.append(list(alteracao.values()))

    # * Definindo intervalo da tabela
    numero_linhas = len(alteracoes) + 1
    tabela_ref = f"A1:E{numero_linhas}"

    # Criando a tabela
    tabela = Table(displayName=nome_planilha.replace(" ", "_"), ref=tabela_ref)

    # Adicionando estilo à tabela
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tabela.tableStyleInfo = style

    # Definindo bordas
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in nova_planilha.iter_rows(
        min_row=1, max_row=numero_linhas, min_col=1, max_col=5
    ):
        for cell in row:
            cell.border = borda

    # Adicionando a tabela na planilha
    nova_planilha.add_table(tabela)

    # Definindo tamanho da linha 1 (cabeçalho)
    nova_planilha.row_dimensions[1].height = 50

    # Alterando Largura das colunas
    tamanhos = [50, 20, 30, 50, 50]
    for coluna, largura in enumerate(tamanhos, start=1):
        nova_planilha.column_dimensions[
            nova_planilha.cell(row=1, column=coluna).column_letter
        ].width = largura


#! Carregando o DataFrame com os boletins mensais:
dataFrame = "planilhas/Extrato Anual.xlsx"
df = load_workbook(dataFrame)

# ? Criando Dicionários para definir as planilhas por meses.
meses = {
    "Maio": df["Maio(212)"],
    "Junho": df["Junho(318)"],
    "Julho": df["Julho(384)"],
    "Agosto": df["Agosto(413)"],
    "Setembro": df["Setembro(485)"],
}

# ? Definindo as colunas da planilha
colunas = {
    "Nome Servidor": "A",
    "Siape": "B",
    "Lotação": "C",
    "Equipe de Trabalho Remoto": "D",
    "Modalidade": "E",
    "Regime de Execução": "F",
}

#! Criando um dicionário para cada mês
dicionario_mensal = {}
for mes, mes_nome in meses.items():
    dicionario_mensal[mes] = criar_dicionario_do_mes(mes_nome)

#! Encontrando as alterações entre os meses
alteracoes = {}
desligado = {}
# * Criando numeração para percorrer os meses
for mes, mes_atual in enumerate(list(meses.keys())[1:], start=1):
    mes_anterior = list(meses.keys())[mes - 1]

    # * Comparando os meses
    alteracoes[mes_atual] = comparar_meses(
        dicionario_mensal[mes_atual], dicionario_mensal[mes_anterior]
    )

    #! Criando nova planilha com os dados que foram alterados
    # * Definindo nome da planilha
    nome_planilha_alteracoes = f"Alterações de {mes_atual}"
    # * Criando cabeçalho
    cabecalho = ["NOME DO AGENTE PÚBLICO", "SIAPE", "ALTERAÇÃO", "DE:", "PARA:"]
    cria_planilha_com_alteracoes(
        df, nome_planilha_alteracoes, alteracoes[mes_atual], cabecalho, mes_atual
    )

#! Salvando planilha com as novas planilhas
caminho_final = "planilhas/Extrato Anual com alterações.xlsx"

df.save(caminho_final)

print(f"Planilha salva com sucesso em {caminho_final}")
